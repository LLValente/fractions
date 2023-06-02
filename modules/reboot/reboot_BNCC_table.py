from openpyxl import Workbook, load_workbook
import pyodbc
import datetime
import os


def main():

    main_path = "C:\\Users\\lucas\\VSCodeProjects\\TCC_Estácio_git"

    file_path = os.path.join(main_path, 'data\\backup')

    wb_ef = load_workbook(filename = os.path.join(file_path, 'BNCC_Ensino Fundamental Simple.xlsx'))
    wb_em = load_workbook(filename = os.path.join(file_path, 'BNCC_Ensino_Medio Simple.xlsx'))

    ws_ef = wb_ef['Planilha1']
    ws_em = wb_em['Planilha1']

    connection_info =    ("Driver={SQL Server};"
                        "Server=DESKTOP-EKI2AGF\SQLEXPRESS;"
                        "Database=SQL_Exams_DB;")

    connection = pyodbc.connect(connection_info)

    cursor = connection.cursor()

    command = f"""DELETE FROM BNCC"""

    cursor.execute(command)

    for i in range(2, ws_ef.max_row + 1):
        
        id_competence = ws_ef[f'C{i}'].value[1:9]

        if id_competence == None:

            break

        school_year = 'EF0' + ws_ef[f'A{i}'].value[0]
        theme = ws_ef[f'B{i}'].value
        competence_description = ws_ef[f'C{i}'].value[11:]
        register_date = str(datetime.datetime.now())[:-7]
        degister_user = os.getlogin()

        command = f"""INSERT INTO BNCC(id_competence, school_year, theme, competence_description, register_date, register_user) VALUES ('{id_competence}', '{school_year}', '{theme}', '{competence_description}', '{register_date}', '{degister_user}')"""
        
        cursor.execute(command)
        cursor.commit()

    for i in range(2, ws_ef.max_row):
        
        id_competence = ws_em[f'B{i}'].value

        if id_competence == None:

            break

        school_year = 'EM*'
        theme = 'NULL'
        competence_description = ws_em[f'C{i}'].value
        register_date = str(datetime.datetime.now())[:-7]
        degister_user = os.getlogin()

        command = f"""INSERT INTO BNCC(id_competence, school_year, theme, competence_description, register_date, register_user) VALUES ('{id_competence}', '{school_year}', '{theme}', '{competence_description}', '{register_date}', '{degister_user}')"""
        
        cursor.execute(command)
        cursor.commit()

    command = f"""DELETE FROM BNCC
        WHERE id_competence = 'None'"""

    cursor.execute(command)
    cursor.commit()

    command = f"""ALTER TABLE Questions ADD CONSTRAINT fk_id_competence FOREIGN KEY (id_competence) REFERENCES BNCC (id_competence)
ALTER TABLE Exams ADD CONSTRAINT fk_id_question FOREIGN KEY (id_question) REFERENCES Questions (id_question)
ALTER TABLE Exams ADD CONSTRAINT fk_id_competence2 FOREIGN KEY (id_competence) REFERENCES BNCC (id_competence)"""

    cursor.execute(command)
    cursor.commit()

    connection.close()

main()