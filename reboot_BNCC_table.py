from openpyxl import Workbook, load_workbook
import pyodbc
import datetime
import os


def main():

    wb_ef = load_workbook(filename = 'BNCC_Ensino Fundamental Simple.xlsx')
    wb_em = load_workbook(filename = 'BNCC_Ensino_Medio Simple.xlsx')

    ws_ef = wb_ef['Planilha1']
    ws_em = wb_em['Planilha1']

    conection_info =    ("Driver={SQL Server};"
                        "Server=DESKTOP-EKI2AGF\SQLEXPRESS;"
                        "Database=SQL_Exams_DB;")

    conection = pyodbc.connect(conection_info)

    cursor = conection.cursor()

    for i in range(2, ws_ef.max_row + 1):
        
        id_competence = ws_ef[f'C{i}'].value[1:9]
        school_year = 'EF0' + ws_ef[f'A{i}'].value[0]
        theme = ws_ef[f'B{i}'].value
        competence_description = ws_ef[f'C{i}'].value[11:]
        register_date = str(datetime.datetime.now())[:-7]
        degister_user = os.getlogin()

        command = f"""INSERT INTO BNCC(id_competence, school_year, theme, competence_description, register_date, register_user) VALUES ('{id_competence}', '{school_year}', '{theme}', '{competence_description}', '{register_date}', '{degister_user}')"""
        
        cursor.execute(command)
        cursor.commit()

    for i in range(2, ws_em.max_row - 1):
        
        id_competence = ws_em[f'B{i}'].value
        school_year = 'EM*'
        theme = 'NULL'
        competence_description = ws_em[f'C{i}'].value
        register_date = str(datetime.datetime.now())[:-7]
        degister_user = os.getlogin()

        command = f"""INSERT INTO BNCC(id_competence, school_year, theme, competence_description, register_date, register_user) VALUES ('{id_competence}', '{school_year}', '{theme}', '{competence_description}', '{register_date}', '{degister_user}')"""
        
        cursor.execute(command)
        cursor.commit()

    command = f"""DELETE FROM BNCC
        WHERE id_competence = 'None'"""

    cursor.execute(command)
    cursor.commit()

main()